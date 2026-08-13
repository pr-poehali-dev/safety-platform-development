import json
import os
import io
import time
import uuid
import base64
import datetime
import boto3
import psycopg2
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook

SCHEMA = "t_p5901577_safety_platform_deve"

CORS = {
    "Access-Control-Allow-Origin": "*",
    "Access-Control-Allow-Methods": "POST, OPTIONS",
    "Access-Control-Allow-Headers": "Content-Type",
}

# Столбцы файла экспорта/импорта, в этом порядке
COL_NUMBER = 0
COL_DATE = 1
COL_OBJECT = 2
COL_CONTRACTOR = 3
COL_INSPECTOR = 4
COL_REPRESENTATIVE = 5
COL_RESPONSIBLE = 6
COL_STATUS = 7
COL_REMARK_NO = 8
COL_PLACE = 9
COL_DESCRIPTION = 10
COL_NORM_REF = 11
COL_DEADLINE = 12
COL_REMARK_STATUS = 13
COL_PHOTO = 14  # 0-based индекс столбца "Фото"

VALID_STATUSES = {"Черновик", "В работе", "Устранено", "Просрочено"}


def get_conn():
    return psycopg2.connect(os.environ["DATABASE_URL"])


def ok(data):
    return {"statusCode": 200, "headers": {**CORS, "Content-Type": "application/json"}, "body": json.dumps(data, ensure_ascii=False)}


def err(msg, code=400):
    return {"statusCode": code, "headers": {**CORS, "Content-Type": "application/json"}, "body": json.dumps({"error": msg})}


MAX_UPLOAD_WORKERS = 40


def s3_client():
    return boto3.client(
        "s3",
        endpoint_url="https://bucket.poehali.dev",
        aws_access_key_id=os.environ["AWS_ACCESS_KEY_ID"],
        aws_secret_access_key=os.environ["AWS_SECRET_ACCESS_KEY"],
        config=boto3.session.Config(max_pool_connections=MAX_UPLOAD_WORKERS),
    )


def upload_photo_bytes(s3, data: bytes) -> str:
    file_key = f"prescription-photos/{uuid.uuid4()}.jpeg"
    s3.put_object(Bucket="files", Key=file_key, Body=data, ContentType="image/jpeg")
    return f"https://cdn.poehali.dev/projects/{os.environ['AWS_ACCESS_KEY_ID']}/bucket/{file_key}"


def upload_photos_parallel(s3, photo_tasks):
    """photo_tasks: список (key, bytes). Возвращает {key: url}, загружая параллельно."""
    results = {}
    if not photo_tasks:
        return results
    with ThreadPoolExecutor(max_workers=min(MAX_UPLOAD_WORKERS, len(photo_tasks))) as executor:
        future_map = {executor.submit(upload_photo_bytes, s3, data): key for key, data in photo_tasks}
        for future in as_completed(future_map):
            key = future_map[future]
            try:
                results[key] = future.result()
            except Exception:
                pass
    return results


def norm(v):
    if v is None:
        return ""
    return str(v).strip()


def parse_workbook(xlsx_bytes: bytes):
    """Парсит xlsx (формат экспорта предписаний) и возвращает список предписаний с замечаниями.
    Каждое изображение из столбца "Фото" привязывается к строке-замечанию по номеру строки листа."""
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active

    # Собираем изображения по 1-based номеру строки (anchor.row — 0-based, поэтому +1)
    images_by_row: dict = {}
    for img in ws._images:
        try:
            row0 = img.anchor._from.row
            col0 = img.anchor._from.col
        except Exception:
            continue
        if col0 != COL_PHOTO:
            continue
        row1 = row0 + 1
        try:
            data = img._data()
        except Exception:
            continue
        images_by_row.setdefault(row1, []).append(data)

    rows = list(ws.iter_rows(min_row=2, values_only=True))

    prescriptions = []
    current = None

    for i, row in enumerate(rows):
        row_idx = i + 2  # 1-based номер строки листа
        number = norm(row[COL_NUMBER]) if len(row) > COL_NUMBER else ""
        remark_no = row[COL_REMARK_NO] if len(row) > COL_REMARK_NO else None
        description = norm(row[COL_DESCRIPTION]) if len(row) > COL_DESCRIPTION else ""

        # Полностью пустая строка — пропускаем
        if not number and not remark_no and not description:
            continue

        if number:
            current = {
                "number": number,
                "date": norm(row[COL_DATE]),
                "object": norm(row[COL_OBJECT]),
                "contractor": norm(row[COL_CONTRACTOR]),
                "inspector": norm(row[COL_INSPECTOR]),
                "representative": norm(row[COL_REPRESENTATIVE]),
                "responsible": norm(row[COL_RESPONSIBLE]),
                "status": norm(row[COL_STATUS]),
                "remarks": [],
            }
            prescriptions.append(current)

        if current is None:
            continue

        if remark_no or description:
            status = norm(row[COL_REMARK_STATUS])
            if status not in VALID_STATUSES:
                status = "В работе"
            photos_bytes = images_by_row.get(row_idx, [])
            current["remarks"].append({
                "place": norm(row[COL_PLACE]),
                "description": description,
                "normRef": norm(row[COL_NORM_REF]),
                "deadline": norm(row[COL_DEADLINE]),
                "status": status,
                "_photos_bytes": photos_bytes,
            })

    return prescriptions


def handler(event: dict, context) -> dict:
    """Импорт предписаний из Excel-файла (формат совпадает с экспортом). Два режима:
    action=preview — распознаёт файл и возвращает сводку без записи в БД;
    action=confirm — сохраняет ранее загруженный файл (по file_key) в базу."""
    if event.get("httpMethod") == "OPTIONS":
        return {"statusCode": 200, "headers": CORS, "body": ""}

    if event.get("httpMethod") != "POST":
        return {"statusCode": 405, "headers": CORS, "body": json.dumps({"error": "method not allowed"})}

    body = json.loads(event.get("body") or "{}")
    action = body.get("action", "preview")

    s3 = s3_client()

    if action == "preview":
        data_url = body.get("fileDataUrl", "")
        if "," not in data_url:
            return err("fileDataUrl is required")
        _, encoded = data_url.split(",", 1)
        xlsx_bytes = base64.b64decode(encoded)

        try:
            prescriptions = parse_workbook(xlsx_bytes)
        except Exception as e:
            return err(f"Не удалось прочитать файл: {e}")

        if not prescriptions:
            return err("В файле не найдено ни одного предписания. Проверьте, что структура файла соответствует формату экспорта.")

        # Проверяем номера предписаний из файла на совпадение с уже существующими в системе
        file_numbers = {p["number"] for p in prescriptions if p["number"]}
        duplicate_numbers = set()
        if file_numbers:
            conn = get_conn()
            cur = conn.cursor()
            try:
                cur.execute(
                    f"SELECT DISTINCT number FROM {SCHEMA}.prescriptions WHERE number = ANY(%s)",
                    (list(file_numbers),)
                )
                duplicate_numbers = {r[0] for r in cur.fetchall()}
            finally:
                cur.close()
                conn.close()

        # Сохраняем оригинальный файл во временное хранилище S3, чтобы не гонять его повторно при подтверждении
        file_key = f"imports/prescriptions-{uuid.uuid4()}.xlsx"
        s3.put_object(
            Bucket="files", Key=file_key, Body=xlsx_bytes,
            ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        remarks_count = sum(len(p["remarks"]) for p in prescriptions)
        photos_count = sum(len(r["_photos_bytes"]) for p in prescriptions for r in p["remarks"])

        preview = [
            {
                "number": p["number"], "date": p["date"], "object": p["object"],
                "contractor": p["contractor"], "remarksCount": len(p["remarks"]),
                "isDuplicate": p["number"] in duplicate_numbers,
            }
            for p in prescriptions
        ]

        return ok({
            "fileKey": file_key,
            "prescriptionsCount": len(prescriptions),
            "remarksCount": remarks_count,
            "photosCount": photos_count,
            "duplicateNumbers": sorted(duplicate_numbers),
            "preview": preview,
        })

    if action == "confirm":
        file_key = body.get("fileKey")
        created_by = body.get("createdBy", "")
        created_by_name = body.get("createdByName", "")
        # "update" — предписания с совпадающим номером обновляют существующую запись;
        # "create" (по умолчанию) — все предписания импортируются как новые записи
        duplicate_mode = body.get("duplicateMode", "create")
        if not file_key:
            return err("fileKey is required")

        obj = s3.get_object(Bucket="files", Key=file_key)
        xlsx_bytes = obj["Body"].read()

        try:
            prescriptions = parse_workbook(xlsx_bytes)
        except Exception as e:
            return err(f"Не удалось прочитать файл: {e}")

        # Если выбрано обновление — находим id существующих предписаний по номеру
        existing_id_by_number = {}
        if duplicate_mode == "update":
            file_numbers = list({p["number"] for p in prescriptions if p["number"]})
            if file_numbers:
                conn0 = get_conn()
                cur0 = conn0.cursor()
                try:
                    cur0.execute(
                        f"SELECT number, id FROM {SCHEMA}.prescriptions WHERE number = ANY(%s)",
                        (file_numbers,)
                    )
                    # При нескольких предписаниях с одинаковым номером берём самое раннее (первое найденное)
                    for num, pid in cur0.fetchall():
                        existing_id_by_number.setdefault(num, pid)
                finally:
                    cur0.close()
                    conn0.close()

        # Присваиваем id заранее и собираем все фото в один пакет для параллельной загрузки
        for p in prescriptions:
            existing_pid = existing_id_by_number.get(p["number"]) if p["number"] else None
            p["_pid"] = existing_pid or f"{int(time.time() * 1000)}{uuid.uuid4().hex[:6]}"
            p["_is_update"] = existing_pid is not None
            for r in p["remarks"]:
                r["_rid"] = f"{int(time.time() * 1000)}{uuid.uuid4().hex[:6]}"

        photo_tasks = []
        for p in prescriptions:
            for r in p["remarks"]:
                for idx, photo_bytes in enumerate(r["_photos_bytes"]):
                    photo_tasks.append(((r["_rid"], idx), photo_bytes))

        uploaded = upload_photos_parallel(s3, photo_tasks)

        conn = get_conn()
        cur = conn.cursor()
        try:
            imported_count = 0
            updated_count = 0
            remarks_count = 0
            for p in prescriptions:
                pid = p["_pid"]
                if p["_is_update"]:
                    log_entry = json.dumps([{
                        "date": datetime.datetime.now(datetime.timezone.utc).isoformat(),
                        "adminLogin": created_by,
                        "adminName": created_by_name,
                    }], ensure_ascii=False)
                    cur.execute(
                        f"UPDATE {SCHEMA}.prescriptions SET date=%s, object=%s, contractor=%s, inspector=%s, "
                        f"representative=%s, responsible=%s, import_log = import_log || %s::jsonb WHERE id=%s",
                        (p["date"], p["object"], p["contractor"], p["inspector"],
                         p["representative"], p["responsible"], log_entry, pid)
                    )
                    cur.execute(f"DELETE FROM {SCHEMA}.remarks WHERE prescription_id = %s", (pid,))
                    updated_count += 1
                else:
                    cur.execute(
                        f"INSERT INTO {SCHEMA}.prescriptions "
                        f"(id, number, date, object, contractor, inspector, representative, responsible, reply_email, report_deadline, comments, created_by) "
                        f"VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                        (pid, p["number"] or f"IMPORT-{pid}", p["date"], p["object"], p["contractor"],
                         p["inspector"], p["representative"], p["responsible"], "", "", "[]", created_by)
                    )
                    imported_count += 1
                for i, r in enumerate(p["remarks"]):
                    photo_urls = [uploaded[(r["_rid"], idx)] for idx in range(len(r["_photos_bytes"])) if (r["_rid"], idx) in uploaded]
                    cur.execute(
                        f"INSERT INTO {SCHEMA}.remarks "
                        f"(id, prescription_id, place, category, description, norm_ref, deadline, status, sort_order, photos) "
                        f"VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                        (r["_rid"], pid, r["place"], "", r["description"], r["normRef"], r["deadline"], r["status"], i,
                         json.dumps(photo_urls, ensure_ascii=False))
                    )
                    remarks_count += 1
            conn.commit()
        finally:
            cur.close()
            conn.close()

        # Импортированный временный файл больше не нужен
        try:
            s3.delete_object(Bucket="files", Key=file_key)
        except Exception:
            pass

        return ok({"ok": True, "prescriptionsCount": imported_count, "updatedCount": updated_count, "remarksCount": remarks_count})

    if action == "cancel":
        file_key = body.get("fileKey")
        if file_key:
            try:
                s3.delete_object(Bucket="files", Key=file_key)
            except Exception:
                pass
        return ok({"ok": True})

    return err("unknown action")