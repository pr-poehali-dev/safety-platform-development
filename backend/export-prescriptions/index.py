import json
import os
import io
import uuid
import boto3
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from PIL import Image as PILImage

CORS = {
    "Access-Control-Allow-Origin": "*",
    "Access-Control-Allow-Methods": "POST, OPTIONS",
    "Access-Control-Allow-Headers": "Content-Type",
}

COLUMNS = [
    "Номер", "Дата", "Объект", "Подрядчик", "Инспектор", "В присутствии", "Ответственный",
    "Статус предписания", "Замечание №", "Место нарушения", "Описание нарушения", "НПА/ЛНА",
    "Срок устранения", "Статус замечания", "Фото",
]
COL_WIDTHS = [10, 12, 24, 24, 22, 22, 22, 16, 10, 20, 40, 22, 14, 14, 42]

THUMB_SIZE = 90
ROW_HEIGHT_WITH_PHOTOS = 74
ROW_HEIGHT_PLAIN = 16
FETCH_TIMEOUT = 3
MAX_WORKERS = 64

_adapter = requests.adapters.HTTPAdapter(pool_connections=MAX_WORKERS, pool_maxsize=MAX_WORKERS)
_session = requests.Session()
_session.mount("https://", _adapter)
_session.mount("http://", _adapter)


def fetch_thumbnail(url: str):
    """Скачивает фото по URL и уменьшает его до миниатюры для вставки в Excel."""
    try:
        resp = _session.get(url, timeout=FETCH_TIMEOUT)
        if resp.status_code != 200:
            return None
        img = PILImage.open(io.BytesIO(resp.content))
        if img.mode not in ("RGB", "RGBA"):
            img = img.convert("RGB")
        img.thumbnail((THUMB_SIZE, THUMB_SIZE))
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=70)
        buf.seek(0)
        return buf, img.width, img.height
    except Exception:
        return None


def fetch_all_thumbnails(tasks):
    """Параллельно скачивает миниатюры для списка (row_idx, url) и возвращает {(row_idx, url): (buf, w, h)}."""
    results = {}
    if not tasks:
        return results
    with ThreadPoolExecutor(max_workers=min(MAX_WORKERS, len(tasks))) as executor:
        future_map = {executor.submit(fetch_thumbnail, url): (row_idx, url) for row_idx, url in tasks}
        for future in as_completed(future_map):
            key = future_map[future]
            try:
                result = future.result()
            except Exception:
                result = None
            if result:
                results[key] = result
    return results


def add_photos_to_cell(ws, sheet_row_idx: int, col_idx: int, photo_urls, thumbnails):
    """Вставляет уже скачанные миниатюры фотографий в ячейку.
    sheet_row_idx — 1-based номер строки листа (используется как ключ в thumbnails и для анкера, где row задаётся 0-based)."""
    x_offset_px = 2
    anchor_row0 = sheet_row_idx - 1
    for url in photo_urls:
        result = thumbnails.get((sheet_row_idx, url))
        if not result:
            continue
        buf, w, h = result
        xl_img = XLImage(buf)
        marker = AnchorMarker(col=col_idx, colOff=pixels_to_EMU(x_offset_px), row=anchor_row0, rowOff=pixels_to_EMU(2))
        size = XDRPositiveSize2D(pixels_to_EMU(w), pixels_to_EMU(h))
        xl_img.anchor = OneCellAnchor(_from=marker, ext=size)
        ws.add_image(xl_img)
        x_offset_px += w + 6


def handler(event: dict, context) -> dict:
    """Экспорт предписаний в Excel со встроенными фотографиями нарушений в столбце "Фото"."""
    if event.get("httpMethod") == "OPTIONS":
        return {"statusCode": 200, "headers": CORS, "body": ""}

    if event.get("httpMethod") != "POST":
        return {"statusCode": 405, "headers": CORS, "body": json.dumps({"error": "method not allowed"})}

    body = json.loads(event.get("body") or "{}")
    prescriptions = body.get("prescriptions", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Предписания"

    for i, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=i, value=col_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for i, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    photo_col_idx0 = len(COLUMNS) - 1  # 0-based индекс столбца "Фото"

    # Первый проход: собираем строки таблицы и список фото для скачивания
    data_rows = []  # (values, photos)
    for p in prescriptions:
        remarks = p.get("remarks", [])
        status = p.get("status", "")
        if not remarks:
            values = [
                p.get("number", ""), p.get("date", ""), p.get("object", ""), p.get("contractor", ""),
                p.get("inspector", ""), p.get("representative", ""), p.get("responsible", ""),
                status, "", "", "", "", "", "", "",
            ]
            data_rows.append((values, []))
            continue

        for idx, r in enumerate(remarks):
            values = [
                p.get("number", "") if idx == 0 else "",
                p.get("date", "") if idx == 0 else "",
                p.get("object", "") if idx == 0 else "",
                p.get("contractor", "") if idx == 0 else "",
                p.get("inspector", "") if idx == 0 else "",
                p.get("representative", "") if idx == 0 else "",
                p.get("responsible", "") if idx == 0 else "",
                status if idx == 0 else "",
                idx + 1,
                r.get("place", ""),
                r.get("description", ""),
                r.get("normRef", ""),
                r.get("deadline", ""),
                r.get("status", ""),
                "",
            ]
            data_rows.append((values, r.get("photos") or []))

    # Параллельно скачиваем все миниатюры разом (вместо последовательных запросов)
    fetch_tasks = []
    for i, (_, photos) in enumerate(data_rows):
        row_idx = i + 2
        for url in photos:
            fetch_tasks.append((row_idx, url))
    thumbnails = fetch_all_thumbnails(fetch_tasks)

    # Второй проход: заполняем лист данными и вставляем уже скачанные миниатюры
    for i, (values, photos) in enumerate(data_rows):
        row_idx = i + 2
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=c, value=v)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        if photos:
            ws.row_dimensions[row_idx].height = ROW_HEIGHT_WITH_PHOTOS
            add_photos_to_cell(ws, row_idx, photo_col_idx0, photos, thumbnails)
        else:
            ws.row_dimensions[row_idx].height = ROW_HEIGHT_PLAIN

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    file_key = f"exports/prescriptions-{uuid.uuid4()}.xlsx"
    s3 = boto3.client(
        "s3",
        endpoint_url="https://bucket.poehali.dev",
        aws_access_key_id=os.environ["AWS_ACCESS_KEY_ID"],
        aws_secret_access_key=os.environ["AWS_SECRET_ACCESS_KEY"],
    )
    s3.put_object(
        Bucket="files",
        Key=file_key,
        Body=buf.getvalue(),
        ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    cdn_url = f"https://cdn.poehali.dev/projects/{os.environ['AWS_ACCESS_KEY_ID']}/bucket/{file_key}"

    return {
        "statusCode": 200,
        "headers": {**CORS, "Content-Type": "application/json"},
        "body": json.dumps({"url": cdn_url}),
    }